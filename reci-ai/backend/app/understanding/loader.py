import json
from typing import List, Dict, Any

class Loader:
    @staticmethod
    def load_json_bytes(json_bytes: bytes) -> List[Dict[str, Any]]:
        """
        Parses raw JSON bytes stream into list of candidate dictionaries.
        """
        try:
            data = json.loads(json_bytes.decode("utf-8"))
            if isinstance(data, dict):
                # If wrapped inside key "candidates", extract it
                if "candidates" in data:
                    return data["candidates"]
                return [data]
            return data
        except Exception as e:
            raise ValueError(f"Failed to parse JSON: {str(e)}")

    @staticmethod
    def _parse_list_field(val: Any) -> List[str]:
        if not val:
            return []
        if isinstance(val, list):
            return [str(x).strip() for x in val if x]
        val_str = str(val).strip()
        if val_str.startswith("[") and val_str.endswith("]"):
            try:
                parsed = json.loads(val_str)
                if isinstance(parsed, list):
                    return [str(x).strip() for x in parsed if x]
            except Exception:
                pass
        # Split by comma or semicolon
        delimiters = [",", ";"]
        for delim in delimiters:
            if delim in val_str:
                return [x.strip() for x in val_str.split(delim) if x.strip()]
        return [val_str] if val_str else []

    @staticmethod
    def _parse_json_or_text_list(val: Any, fallback_key: str = "description") -> List[Dict[str, Any]]:
        if not val:
            return []
        if isinstance(val, list):
            return [x for x in val if isinstance(x, dict)]
        val_str = str(val).strip()
        if val_str.startswith("[") and val_str.endswith("]"):
            try:
                parsed = json.loads(val_str)
                if isinstance(parsed, list):
                    result = []
                    for x in parsed:
                        if isinstance(x, dict):
                            result.append(x)
                        elif isinstance(x, str):
                            result.append({fallback_key: x})
                    return result
            except Exception:
                pass
        if val_str.startswith("{") and val_str.endswith("}"):
            try:
                parsed = json.loads(val_str)
                if isinstance(parsed, dict):
                    return [parsed]
            except Exception:
                pass
        return [{fallback_key: val_str}]

    @staticmethod
    def _row_to_candidate(row_dict: Dict[str, Any]) -> Dict[str, Any]:
        candidate = {}
        
        # Normalize keys to lowercase with no spaces/underscores
        norm_row = {
            str(k).lower().replace(" ", "").replace("_", ""): v 
            for k, v in row_dict.items() 
            if v is not None and str(v).strip() != ""
        }
        
        # 1. candidate_id
        candidate_id = ""
        for k in ["candidateid", "id", "candid"]:
            if k in norm_row:
                candidate_id = str(norm_row[k]).strip()
                break
        if not candidate_id:
            name_val = str(norm_row.get("name", "")).strip()
            if name_val:
                candidate_id = "cand-" + name_val.lower().replace(" ", "-")
            else:
                import uuid
                candidate_id = "cand-" + str(uuid.uuid4())[:8]
        candidate["candidate_id"] = candidate_id
        
        # 2. name
        candidate["name"] = str(norm_row.get("name", norm_row.get("fullname", "Unknown Candidate"))).strip()
        
        # 3. skills
        skills_val = norm_row.get("skills", norm_row.get("skillset", ""))
        candidate["skills"] = Loader._parse_list_field(skills_val)
            
        # 4. work_history
        wh_val = norm_row.get("workhistory", norm_row.get("experience", ""))
        candidate["work_history"] = Loader._parse_json_or_text_list(wh_val, fallback_key="description")
            
        # 5. projects
        proj_val = norm_row.get("projects", norm_row.get("portfolio", ""))
        candidate["projects"] = Loader._parse_json_or_text_list(proj_val, fallback_key="description")
            
        # 6. education
        edu_val = norm_row.get("education", norm_row.get("degree", ""))
        candidate["education"] = Loader._parse_json_or_text_list(edu_val, fallback_key="degree")
            
        # 7. certifications
        cert_val = norm_row.get("certifications", norm_row.get("certs", ""))
        candidate["certifications"] = Loader._parse_json_or_text_list(cert_val, fallback_key="certification")
            
        # 8. languages
        lang_val = norm_row.get("languages", norm_row.get("lang", ""))
        candidate["languages"] = Loader._parse_list_field(lang_val)
            
        # 9. activity_metadata
        meta_val = norm_row.get("activitymetadata", norm_row.get("metadata", ""))
        if meta_val:
            try:
                candidate["activity_metadata"] = json.loads(str(meta_val)) if isinstance(meta_val, str) else meta_val
            except Exception:
                candidate["activity_metadata"] = {"raw_value": str(meta_val)}
        else:
            candidate["activity_metadata"] = {}
            
        # 10. experience_years
        import re
        exp_years = 0.0
        for k in ["experienceyears", "years", "exp"]:
            if k in norm_row:
                val = norm_row[k]
                try:
                    match = re.search(r'([\d\.]+)', str(val))
                    if match:
                        exp_years = float(match.group(1))
                        break
                except Exception:
                    pass
        else:
            if "experience" in norm_row:
                val_str = str(norm_row["experience"]).strip()
                if re.match(r'^[\d\.]+(?:\s*(?:years|yr|years\s*exp|yrs|year))?$', val_str, re.IGNORECASE):
                    try:
                        match = re.search(r'([\d\.]+)', val_str)
                        if match:
                            exp_years = float(match.group(1))
                    except Exception:
                        pass
        candidate["experience_years"] = exp_years
            
        return candidate

    @staticmethod
    def load_csv_bytes(csv_bytes: bytes) -> List[Dict[str, Any]]:
        """
        Parses raw CSV bytes stream into list of candidate dictionaries.
        """
        import csv
        import io
        try:
            content = csv_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            content = csv_bytes.decode("latin-1")
            
        reader = csv.DictReader(io.StringIO(content))
        candidates = []
        for row in reader:
            candidates.append(Loader._row_to_candidate(row))
        return candidates

    @staticmethod
    def load_excel_bytes(excel_bytes: bytes) -> List[Dict[str, Any]]:
        """
        Parses raw Excel (.xlsx) bytes stream into list of candidate dictionaries.
        """
        import openpyxl
        import io
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            sheet = wb.active
            
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
                
            candidates = []
            for r in range(2, sheet.max_row + 1):
                row_dict = {}
                has_data = False
                for c, header in enumerate(headers, 1):
                    if header is not None:
                        val = sheet.cell(row=r, column=c).value
                        if val is not None and str(val).strip() != "":
                            has_data = True
                        row_dict[header] = val
                if has_data:
                    candidates.append(Loader._row_to_candidate(row_dict))
            return candidates
        except Exception as e:
            raise ValueError(f"Failed to parse Excel: {str(e)}")

    @staticmethod
    def get_sample_candidates() -> List[Dict[str, Any]]:
        """
        Returns a high-fidelity sample candidate dataset for default runs or out-of-the-box UI demonstration.
        """
        return [
            {
                "candidate_id": "cand-01",
                "name": "Sarah Jenkins",
                "skills": ["React", "node", "javascript", "amazon web services", "Postgres"],
                "work_history": [
                    {
                        "company": "Tech Corp",
                        "role": "Software Developer",
                        "start_date": "2021-01",
                        "end_date": "2023-05",
                        "description": "Built standard application features using React and Node.js."
                    },
                    {
                        "company": "Innovate LLC",
                        "role": "Senior Full Stack Engineer",
                        "start_date": "2023-06",
                        "end_date": "Present",
                        "description": "Leading team developers. Upgrading platform architecture to AWS microservices."
                    }
                ],
                "projects": [
                    {
                        "name": "SaaS Platform Migration",
                        "description": "Migrated legacy monolith backend to node/react on AWS cloud for a fintech banking application.",
                        "technologies": ["React", "Node", "AWS"],
                        "role": "Senior Architect",
                        "duration_months": 12
                    }
                ],
                "education": [
                    {
                        "degree": "B.S.",
                        "major": "Computer Science",
                        "institution": "State University",
                        "year": 2020
                    }
                ],
                "certifications": [
                    {
                        "certification": "AWS Certified Solutions Architect",
                        "provider": "",
                        "year": 2022,
                        "category": "Cloud"
                    }
                ],
                "languages": ["English", "Spanish"],
                "activity_metadata": {
                    "github_contributions": 142,
                    "platform_logins": 12
                }
            },
            {
                "candidate_id": "cand-02",
                "name": "David Kojo",
                "skills": ["Python", "PyTorch", "Kubernetes", "Docker", "gcp"],
                "work_history": [
                    {
                        "company": "DeepMind Co",
                        "role": "ML Engineer",
                        "start_date": "2019-06",
                        "end_date": "Present",
                        "description": "Designing ML pipelines and scaling models on Kubernetes cluster."
                    }
                ],
                "projects": [
                    {
                        "name": "Recommendation Engine",
                        "description": "Trained custom collaborative filter models in PyTorch.",
                        "technologies": ["PyTorch", "Python", "Kubernetes"],
                        "role": "ML Engineer",
                        "duration_months": 24
                    }
                ],
                "education": [
                    {
                        "degree": "Masters",
                        "major": "Data Science",
                        "institution": "Tech Institute",
                        "year": 2019
                    }
                ],
                "certifications": [],
                "languages": ["English", "Twi"],
                "activity_metadata": {
                    "github_contributions": 48
                }
            }
        ]

    @staticmethod
    def get_candidate_schema() -> Dict[str, Any]:
        """
        Returns JSON schema representation of candidates for UI downloads or reference check.
        """
        return {
            "$schema": "http://json-schema.org/draft-07/schema#",
            "title": "CandidateDataset",
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "candidate_id": {"type": "string"},
                    "name": {"type": "string"},
                    "skills": {"type": "array", "items": {"type": "string"}},
                    "work_history": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "company": {"type": "string"},
                                "role": {"type": "string"},
                                "start_date": {"type": "string"},
                                "end_date": {"type": "string"},
                                "description": {"type": "string"}
                            },
                            "required": ["company", "role", "start_date"]
                        }
                    },
                    "projects": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "description": {"type": "string"},
                                "technologies": {"type": "array", "items": {"type": "string"}},
                                "role": {"type": "string"},
                                "duration_months": {"type": "integer"}
                            },
                            "required": ["name", "technologies"]
                        }
                    },
                    "education": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "degree": {"type": "string"},
                                "major": {"type": "string"},
                                "institution": {"type": "string"},
                                "year": {"type": "integer"}
                            }
                        }
                    },
                    "certifications": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "certification": {"type": "string"},
                                "provider": {"type": "string"},
                                "year": {"type": "integer"}
                            }
                        }
                    },
                    "languages": {"type": "array", "items": {"type": "string"}},
                    "activity_metadata": {"type": "object"}
                },
                "required": ["candidate_id", "name"]
            }
        }
